import os
import datetime
from openpyxl import Workbook, load_workbook

class Student:
    def __init__(self, name, dob, gender, address, math_score, literature_score, english_score, social_scores=None, natural_scores=None):
        self.name = name
        self.dob = dob
        self.gender = gender
        self.address = address
        self.math_score = math_score
        self.literature_score = literature_score
        self.english_score = english_score
        self.social_scores = social_scores if social_scores else []
        self.natural_scores = natural_scores if natural_scores else []

    def calculate_age(self):
        today = datetime.date.today()
        birth_date = datetime.datetime.strptime(self.dob, "%Y-%m-%d").date()
        age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))
        return age

    def calculate_average_score(self):
        total_scores = sum([self.math_score, self.literature_score, self.english_score] + self.social_scores + self.natural_scores)
        num_subjects = 3 + len(self.social_scores) + len(self.natural_scores)
        return total_scores / num_subjects

def add_student_to_excel(student):
    if not os.path.exists("students.xlsx"):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Date of Birth", "Gender", "Address", "Math Score", "Literature Score", "English Score", "Social Scores", "Natural Scores"])
        wb.save("students.xlsx")

    wb = load_workbook("students.xlsx")
    ws = wb.active
    ws.append([
        student.name, student.dob, student.gender, student.address, 
        student.math_score, student.literature_score, student.english_score, 
        ",".join(map(str, student.social_scores)), ",".join(map(str, student.natural_scores))
    ])
    wb.save("students.xlsx")

def load_students_from_excel():
    students = []
    if os.path.exists("students.xlsx"):
        wb = load_workbook("students.xlsx")
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, dob, gender, address, math_score, literature_score, english_score, social_scores_str, natural_scores_str = row
            social_scores = [float(score) for score in (social_scores_str.split(",") if social_scores_str else []) if score.strip()]
            natural_scores = [float(score) for score in (natural_scores_str.split(",") if natural_scores_str else []) if score.strip()]
            student = Student(name, dob, gender, address, math_score, literature_score, english_score, social_scores, natural_scores)
            students.append(student)
        wb.close()
    return students

def add_student(students):
    os.system('cls' if os.name == 'nt' else 'clear')  # Xóa màn hình
    name = input("Nhập họ và tên của sinh viên: ")
    dob = input("Nhập ngày tháng năm sinh (YYYY-MM-DD): ")
    gender = input("Nhập giới tính: ")
    address = input("Nhập địa chỉ: ")
    math_score = float(input("Nhập điểm môn Toán: "))
    literature_score = float(input("Nhập điểm môn Văn: "))
    english_score = float(input("Nhập điểm môn Anh: "))
    
    social_scores = []
    natural_scores = []

    student_type = input("Nhập loại học sinh (Xã hội hoặc Tự nhiên): ").lower()
    if student_type == "xã hội":
        social_scores.append(float(input("Nhập điểm môn Sử: ")))
        social_scores.append(float(input("Nhập điểm môn Địa: ")))
        social_scores.append(float(input("Nhập điểm môn GDCD: ")))
    elif student_type == "tự nhiên":
        natural_scores.append(float(input("Nhập điểm môn Lý: ")))
        natural_scores.append(float(input("Nhập điểm môn Hóa: ")))
        natural_scores.append(float(input("Nhập điểm môn Sinh: ")))

    student = Student(name, dob, gender, address, math_score, literature_score, english_score, social_scores, natural_scores)
    students.append(student)
    add_student_to_excel(student)
    print("\nĐã thêm sinh viên vào danh sách.")
    input("Nhấn Enter để quay lại menu...")

def display_students(students):
    os.system('cls' if os.name == 'nt' else 'clear')  # Xóa màn hình
    if not students:
        print("Danh sách sinh viên trống.")
    else:
        print("\nDanh sách sinh viên:")
        print("{:<20} {:<10} {:<10} {:<20} {:<10}".format("Họ và Tên", "Tuổi", "Giới tính", "Địa chỉ", "Điểm TB"))
        print("-" * 70)
        for student in students:
            print("{:<20} {:<10} {:<10} {:<20} {:<10}".format(student.name, student.calculate_age(), student.gender, student.address, student.calculate_average_score()))

def update_student(students):
    os.system('cls' if os.name == 'nt' else 'clear')  # Xóa màn hình
    display_students(students)
    if students:
        try:
            choice = int(input("Nhập số thứ tự của sinh viên cần cập nhật: "))
            if 1 <= choice <= len(students):
                student = students[choice - 1]
                print("\nThông tin hiện tại của sinh viên:")
                display_students([student])  # Hiển thị thông tin sinh viên cần cập nhật
                print("\nNhập thông tin mới cho sinh viên:")
                student.name = input("Nhập họ và tên mới của sinh viên: ")
                student.dob = input("Nhập ngày tháng năm sinh mới (DD/MM/YYYY): ")
                student.gender = input("Nhập giới tính mới: ")
                student.address = input("Nhập địa chỉ mới: ")
                student.math_score = float(input("Nhập điểm môn Toán mới: "))
                student.literature_score = float(input("Nhập điểm môn Văn mới: "))
                student.english_score = float(input("Nhập điểm môn Anh mới: "))
                student.social_scores = []
                student.natural_scores = []
                student_type = input("Nhập loại học sinh mới (Xã hội hoặc Tự nhiên): ").lower()
                if student_type == "xã hội":
                    student.social_scores.append(float(input("Nhập điểm môn Sử mới: ")))
                    student.social_scores.append(float(input("Nhập điểm môn Địa mới: ")))
                    student.social_scores.append(float(input("Nhập điểm môn GDCD mới: ")))
                elif student_type == "tự nhiên":
                    student.natural_scores.append(float(input("Nhập điểm môn Lý mới: ")))
                    student.natural_scores.append(float(input("Nhập điểm môn Hóa mới: ")))
                    student.natural_scores.append(float(input("Nhập điểm môn Sinh mới: ")))
                print("\nĐã cập nhật thông tin cho sinh viên.")
                display_students([student])  # Hiển thị thông tin sinh viên sau khi cập nhật
                save_students_to_excel(students)
            else:
                print("Số thứ tự không hợp lệ.")
        except ValueError:
            print("Vui lòng nhập một số nguyên.")
    else:
        print("Không có sinh viên để cập nhật.")
    input("Nhấn Enter để quay lại menu...")

def delete_student(students):
    os.system('cls' if os.name == 'nt' else 'clear')  # Xóa màn hình
    display_students(students)
    if students:
        try:
            choice = int(input("Nhập số thứ tự của sinh viên cần xóa: "))
            if 1 <= choice <= len(students):
                del students[choice - 1]
                print("Đã xóa sinh viên khỏi danh sách.")
                save_students_to_excel(students)
            else:
                print("Số thứ tự không hợp lệ.")
        except ValueError:
            print("Vui lòng nhập một số nguyên.")
    else:
        print("Không có sinh viên để xóa.")
    input("Nhấn Enter để quay lại menu...")

def save_students_to_excel(students):
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Date of Birth", "Gender", "Address", "Math Score", "Literature Score", "English Score", "Social Scores", "Natural Scores"])
    for student in students:
        ws.append([
            student.name, student.dob, student.gender, student.address, 
            student.math_score, student.literature_score, student.english_score, 
            ",".join(map(str, student.social_scores)), ",".join(map(str, student.natural_scores))
        ])
    wb.save("students.xlsx")

def main():
    students = load_students_from_excel()
    while True:
        print("\n===== MENU =====")
        print("1. Thêm sinh viên")
        print("2. Hiển thị danh sách sinh viên")
        print("3. Cập nhật thông tin sinh viên")
        print("4. Xóa sinh viên")
        print("5. Thoát")
        choice = input("Chọn chức năng (1/2/3/4/5): ")

        if choice == "1":
            add_student(students)
        elif choice == "2":
            display_students(students)
        elif choice == "3":
            update_student(students)
        elif choice == "4":
            delete_student(students)
        elif choice == "5":
            print("Đã thoát chương trình.")
            break
        else:
            print("Lựa chọn không hợp lệ. Vui lòng chọn lại.")
        input("Nhấn Enter để tiếp tục...")
        os.system('cls' if os.name == 'nt' else 'clear')  # Xóa màn hình

if __name__ == "__main__":
    main()

        